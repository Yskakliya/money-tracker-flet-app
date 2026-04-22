import flet as ft
import mysql.connector
import os
from datetime import datetime


def get_export_view(page: ft.Page, user_name="User"):
    PRIMARY      = "#85bb65"
    RED          = "#ef4444"
    BLUE         = "#3b82f6"
    CARD_BG      = "#0f2d25"
    BG_DARK      = "#091f1a"
    BORDER_COLOR = "#1e4035"
    TEXT_DIM     = "#6b9e7e"

    db_config = {
        "host": "localhost", "user": "root",
        "password": "venera123!@ZX", "database": "money_tracker_v2",
    }

    status_text  = ft.Text("", color=TEXT_DIM, size=13)
    progress_bar = ft.ProgressBar(visible=False, color=PRIMARY, bgcolor=BORDER_COLOR)

    def load_all_data():
        try:
            db = mysql.connector.connect(**db_config)
            cursor = db.cursor(dictionary=True)
            cursor.execute(
                "SELECT date, amount, category FROM user_transactions WHERE user_name=%s ORDER BY date DESC",
                (user_name,),
            )
            tx = cursor.fetchall()
            cursor.execute(
                "SELECT category_name, limit_amount, spent_amount FROM user_budgets WHERE user_name=%s",
                (user_name,),
            )
            budgets = cursor.fetchall()
            db.close()
            return tx, budgets
        except Exception as e:
            print(f"Export load error: {e}")
            return [], []

    def set_status(msg, color=TEXT_DIM):
        status_text.value = msg
        status_text.color = color
        page.update()

    def export_excel(e):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            set_status("Installing openpyxl...", TEXT_DIM)
            os.system("pip install openpyxl --break-system-packages -q")
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        progress_bar.visible = True
        set_status("Preparing data...")

        tx, budgets = load_all_data()

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Transactions"

        header_fill = PatternFill("solid", fgColor="0F2D25")
        header_font = Font(bold=True, color="85BB65")
        thin = Border(
            bottom=Side(style="thin", color="1E4035"),
        )

        headers = ["Date", "Category", "Amount ($)"]
        col_widths = [18, 20, 15]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws1.cell(row=1, column=ci, value=h)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws1.column_dimensions[cell.column_letter].width = w

        for ri, row in enumerate(tx, 2):
            d = row["date"]
            date_str = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)
            ws1.cell(row=ri, column=1, value=date_str)
            ws1.cell(row=ri, column=2, value=row["category"])
            amt_cell = ws1.cell(row=ri, column=3, value=float(row["amount"]))
            amt_cell.number_format = '"$"#,##0.00'
            for ci in range(1, 4):
                ws1.cell(row=ri, column=ci).border = thin

        total_row = len(tx) + 2
        ws1.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, color="FFFFFF")
        total_cell = ws1.cell(row=total_row, column=3,
                              value=sum(float(r["amount"]) for r in tx))
        total_cell.number_format = '"$"#,##0.00'
        total_cell.font = Font(bold=True, color="85BB65")

        ws2 = wb.create_sheet("Budgets")
        b_headers = ["Category", "Limit ($)", "Spent ($)", "Remaining ($)", "% Used"]
        b_widths   = [20, 14, 14, 16, 10]
        for ci, (h, w) in enumerate(zip(b_headers, b_widths), 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws2.column_dimensions[cell.column_letter].width = w

        for ri, b in enumerate(budgets, 2):
            limit = float(b["limit_amount"])
            spent = float(b["spent_amount"])
            rem   = limit - spent
            pct   = (spent / limit * 100) if limit > 0 else 0
            ws2.cell(row=ri, column=1, value=b["category_name"])
            for ci, val, fmt in [
                (2, limit, '"$"#,##0.00'),
                (3, spent, '"$"#,##0.00'),
                (4, rem,   '"$"#,##0.00'),
                (5, round(pct, 1), '0.0"%"'),
            ]:
                c = ws2.cell(row=ri, column=ci, value=val)
                c.number_format = fmt

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        os.makedirs(desktop, exist_ok=True)
        fname = os.path.join(desktop, f"money_tracker_{user_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        wb.save(fname)

        progress_bar.visible = False
        set_status(f"Excel saved to Desktop: {os.path.basename(fname)}", PRIMARY)

    def export_pdf(e):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            set_status("Installing reportlab...", TEXT_DIM)
            os.system("pip install reportlab --break-system-packages -q")
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm

        progress_bar.visible = True
        set_status("Building PDF...")

        tx, budgets = load_all_data()

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        os.makedirs(desktop, exist_ok=True)
        fname = os.path.join(desktop, f"money_tracker_{user_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf")

        doc = SimpleDocTemplate(fname, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        GREEN  = colors.HexColor("#85bb65")
        DARK   = colors.HexColor("#0f2d25")
        GREY   = colors.HexColor("#6b9e7e")
        WHITE  = colors.white
        RED_C  = colors.HexColor("#ef4444")

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", fontSize=20, textColor=GREEN,
                                     fontName="Helvetica-Bold", spaceAfter=6)
        sub_style   = ParagraphStyle("sub",   fontSize=10, textColor=GREY,
                                     fontName="Helvetica",    spaceAfter=12)
        h2_style    = ParagraphStyle("h2",    fontSize=13, textColor=WHITE,
                                     fontName="Helvetica-Bold", spaceBefore=16, spaceAfter=8)

        story = []

        story.append(Paragraph("Money Tracker — Full Export", title_style))
        story.append(Paragraph(
            f"User: {user_name}   |   Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
            sub_style,
        ))
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph("Budget Summary", h2_style))
        b_data = [["Category", "Limit", "Spent", "Remaining", "% Used"]]
        for b in budgets:
            limit = float(b["limit_amount"])
            spent = float(b["spent_amount"])
            rem   = limit - spent
            pct   = (spent / limit * 100) if limit > 0 else 0
            b_data.append([
                b["category_name"],
                f"${limit:,.0f}",
                f"${spent:,.0f}",
                f"${rem:,.0f}",
                f"{pct:.0f}%",
            ])
        
        tl = sum(float(b["limit_amount"]) for b in budgets)
        ts = sum(float(b["spent_amount"]) for b in budgets)
        b_data.append(["TOTAL", f"${tl:,.0f}", f"${ts:,.0f}", f"${tl-ts:,.0f}", ""])

        b_table = Table(b_data, colWidths=[4.5*cm, 3*cm, 3*cm, 3.5*cm, 2.5*cm])
        b_table.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  DARK),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  GREEN),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("TEXTCOLOR",   (0, 1), (-1, -1), WHITE),
            ("BACKGROUND",  (0, -1), (-1, -1), DARK),
            ("TEXTCOLOR",   (0, -1), (-1, -1), GREEN),
            ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.HexColor("#0b2219"), colors.HexColor("#0f2d25")]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#1e4035")),
            ("ALIGN",       (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN",       (0, 0), (0, -1),  "LEFT"),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(b_table)

        story.append(Paragraph("All Transactions", h2_style))
        t_data = [["Date", "Category", "Amount"]]
        for r in tx:
            d = r["date"]
            date_str = d.strftime("%d %b %Y") if hasattr(d, "strftime") else str(d)
            t_data.append([date_str, r["category"], f"${float(r['amount']):,.2f}"])
        total_amt = sum(float(r["amount"]) for r in tx)
        t_data.append(["", "TOTAL", f"${total_amt:,.2f}"])

        t_table = Table(t_data, colWidths=[4*cm, 6*cm, 4*cm])
        t_table.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0),  (-1, 0),  DARK),
            ("TEXTCOLOR",   (0, 0),  (-1, 0),  GREEN),
            ("FONTNAME",    (0, 0),  (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0),  (-1, -1), 9),
            ("TEXTCOLOR",   (0, 1),  (-1, -1), WHITE),
            ("BACKGROUND",  (0, -1), (-1, -1), DARK),
            ("TEXTCOLOR",   (1, -1), (-1, -1), GREEN),
            ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.HexColor("#0b2219"), colors.HexColor("#0f2d25")]),
            ("GRID",        (0, 0),  (-1, -1), 0.5, colors.HexColor("#1e4035")),
            ("ALIGN",       (2, 0),  (2, -1),  "RIGHT"),
            ("TOPPADDING",  (0, 0),  (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t_table)

        doc.build(story)

        progress_bar.visible = False
        set_status(f"PDF saved to Desktop: {os.path.basename(fname)}", PRIMARY)

    # ── UI ────────────────────────────────────────────────────────────────────
    def export_card(icon, title, subtitle, btn_label, btn_color, action):
        return ft.Container(
            content=ft.Row([
                ft.Container(
                    ft.Icon(icon, color=btn_color, size=28),
                    bgcolor=ft.colors.with_opacity(0.15, btn_color),
                    border_radius=16,
                    padding=16,
                    width=60, height=60,
                    alignment=ft.alignment.center,
                ),
                ft.Column([
                    ft.Text(title,    color="white",   size=15, weight="bold"),
                    ft.Text(subtitle, color=TEXT_DIM,  size=12),
                ], spacing=4, expand=True),
                ft.Container(
                    content=ft.Text(btn_label, color="white", size=13, weight="bold"),
                    bgcolor=btn_color,
                    border_radius=12,
                    padding=ft.padding.symmetric(horizontal=20, vertical=10),
                    on_click=action,
                    ink=True,
                ),
            ], spacing=16),
            bgcolor=CARD_BG,
            border=ft.border.all(1, BORDER_COLOR),
            padding=20,
            border_radius=20,
        )

    return ft.Container(
        content=ft.ListView(
            controls=[
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Column([
                                ft.Text("Export Data", size=26, weight="bold", color="white"),
                                ft.Text("Download all your financial data", color=TEXT_DIM, size=12),
                            ], spacing=2),
                        ]),
                        ft.Container(height=1, bgcolor=BORDER_COLOR),

                        ft.Text("Choose export format", color=TEXT_DIM, size=13),

                        export_card(
                            ft.icons.TABLE_CHART_OUTLINED,
                            "Export to Excel (.xlsx)",
                            "All transactions + budget summary in spreadsheet format",
                            "Export",
                            PRIMARY,
                            export_excel,
                        ),
                        export_card(
                            ft.icons.PICTURE_AS_PDF_OUTLINED,
                            "Export to PDF",
                            "Formatted report with all transactions and budgets",
                            "Export",
                            BLUE,
                            export_pdf,
                        ),

                        progress_bar,
                        status_text,

                        # Info box
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.icons.INFO_OUTLINE, color=TEXT_DIM, size=16),
                                ft.Text(
                                    "Files are saved to your Desktop automatically.",
                                    color=TEXT_DIM, size=12,
                                ),
                            ], spacing=10),
                            bgcolor=ft.colors.with_opacity(0.05, "white"),
                            border=ft.border.all(1, BORDER_COLOR),
                            padding=ft.padding.symmetric(horizontal=16, vertical=12),
                            border_radius=12,
                        ),
                    ], spacing=16),
                    padding=ft.padding.symmetric(horizontal=24, vertical=20),
                )
            ],
            expand=True,
            spacing=0,
            padding=0,
        ),
        expand=True,
        bgcolor=BG_DARK,
    )